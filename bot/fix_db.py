from os import listdir, walk

from openpyxl import load_workbook

from src.database import users_col


for filename in listdir('files_for_init/'):
    wb = load_workbook(f'files_for_init/{filename}')
    platform = filename.split('.')[0].lower()
    ws = wb.active
    for i in range(2, 9999):
        if ws[f'A{i}'].value is None:
            break
        
        full_name = f"{ws[f'A{i}'].value} {ws[f'B{i}'].value} {ws[f'C{i}'].value}"
        id_ = ws[f'D{i}'].value
        users_col.delete_one({'code': id_})
    print(f'Удаляю из {filename}')
    
print('Все лишние пользователи удалены!')