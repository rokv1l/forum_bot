from datetime import date

from openpyxl import load_workbook

from src.database import day_program_col


wb = load_workbook(f'files/программа30.xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    date_ = date(year=2022, month=5, day=31).isoformat()
    platform = sheet_name.lower()
    events = []
    for i in range(3, 9999):
        if ws[f'A{i}'].value is None:
            break
        
        time = str(ws[f'A{i}'].value)
        name = ws[f'B{i}'].value
        if ws[f"C{i}"].value and ws[f"C{i}"].value != '\n':
            name += f'\nМесто проведения: {ws[f"C{i}"].value}'
            
        events.append({"event_name": name, "event_time": time})

    day_program_col.insert_one({'date': date_, 'platform': platform, "events": events})
