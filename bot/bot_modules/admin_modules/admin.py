import os
from datetime import datetime

import openpyxl
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, MessageAutoDeleteTimerChanged
from telegram.ext import CallbackQueryHandler, ConversationHandler, CommandHandler, MessageHandler, Filters

import config
from src.database import events_col, users_col
from src.utils import admin_menu, delete_keyboard
from bot_modules.admin_modules.mailing import mailing_handler
from bot_modules.admin_modules.event import create_event_handler
from bot_modules.admin_modules.support_admin import support_admin_handler



def events_data_callback(update, context):
    if 'msg_for_del_keys' in context.user_data:
        delete_keyboard(context, update.effective_chat.id)
    
    events = events_col.find()
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Лист 1'
    sheet.cell(row=1, column=1).value = 'Категория'
    sheet.cell(row=1, column=2).value = 'Название'
    sheet.cell(row=1, column=3).value = 'Описание'
    sheet.cell(row=1, column=4).value = 'Дата и время начала'
    sheet.cell(row=1, column=5).value = 'Подписавшиеся пользователи'
    for i, event in enumerate(events):
        sheet.cell(row=i + 2, column=1).value = event['category']
        sheet.cell(row=i + 2, column=2).value = event['name']
        sheet.cell(row=i + 2, column=3).value = event['desc']
        sheet.cell(row=i + 2, column=4).value = event['dt'].isoformat().replace('T', ' ')
        
        users = [users_col.find_one({'tg_id': user}) for user in event['users']]
        sheet.cell(row=i + 2, column=5).value = ', '.join([f"{user['full_name']} {user['platform']}" for user in users])
        
    path = f'files/events_data_{update.effective_chat.id}_{datetime.now().isoformat().replace(":", "_")}.xlsx'
    wb.save(path)
    wb.close()
    with open(path, 'rb') as file:
        update.effective_chat.send_document(file, filename=file.name)
        
    os.remove(path)
    return admin_menu(update, context)


def back_callback(update, context):
    if 'msg_for_del_keys' in context.user_data:
        delete_keyboard(context, update.effective_chat.id)
    
    update.effective_chat.send_message('Вы вышли из меню администратора')
    return ConversationHandler.END


def unknown(update, context):
    update.effective_chat.send_message(
        'Вы ввели неизвестную команду.\n'\
        'Если что-то пошло не так то воспользуйтесь /exit для сброса диалогов к началу'
    )


def _exit(update, context):
    if 'msg_for_del_keys' in context.user_data:
        delete_keyboard(context, update.effective_chat.id)

    update.effective_chat.send_message('Вы вышли из всех диалогов, для использования бота нажмите /start')
    return ConversationHandler.END


admin_handler = ConversationHandler(
    entry_points=[
        CommandHandler('admin', admin_menu)
    ],
    states={
        config.ADMIN_MAIN_MENU: [
            mailing_handler,
            create_event_handler,
            CallbackQueryHandler(back_callback, pattern=r'Выход'),
            CallbackQueryHandler(events_data_callback, pattern=r'events_data')
            # support_admin_handler
        ],
    },
    fallbacks=[
        CommandHandler('exit', _exit),
        MessageHandler(Filters.all, unknown)
    ]
)
